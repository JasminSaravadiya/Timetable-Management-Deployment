from fastapi import FastAPI, Depends, HTTPException, status, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete as sa_delete
from typing import List, Optional
from datetime import datetime, timedelta, date, time
import time as _time
import logging

import models
import schemas
from database import get_db, engine, Base

logger = logging.getLogger("uvicorn.error")

app = FastAPI(title="Master Timetable Generator API")

# ── GZip compression for all responses ──
app.add_middleware(GZipMiddleware, minimum_size=500)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Slow query logging middleware ──
@app.middleware("http")
async def log_slow_requests(request: Request, call_next):
    start = _time.perf_counter()
    response = await call_next(request)
    elapsed_ms = (_time.perf_counter() - start) * 1000
    if elapsed_ms > 500:
        logger.warning(f"SLOW REQUEST: {request.method} {request.url.path} took {elapsed_ms:.0f}ms")
    return response

# ── In-memory TTL cache ──
_cache: dict[str, tuple[float, any]] = {}
_CACHE_TTL = 30  # seconds

def _cache_get(key: str):
    entry = _cache.get(key)
    if entry and (_time.time() - entry[0]) < _CACHE_TTL:
        return entry[1]
    return None

def _cache_set(key: str, value):
    _cache[key] = (_time.time(), value)

def _cache_invalidate():
    _cache.clear()

@app.on_event("startup")
async def startup():
    async with engine.begin() as conn:
        await conn.run_sync(models.Base.metadata.create_all)

@app.get("/")
def read_root():
    return {"message": "Welcome to Master Timetable API"}

# ═══════════════════════════════════
#  CONFIG
# ═══════════════════════════════════
@app.post("/api/config", response_model=schemas.ConfigOut)
async def create_config(config: schemas.ConfigCreate, db: AsyncSession = Depends(get_db)):
    config_dict = config.model_dump()
    config_dict['breaks'] = [b.model_dump(mode='json') for b in config.breaks]
    db_config = models.TimetableConfig(**config_dict)
    db.add(db_config)
    await db.commit()
    await db.refresh(db_config)
    _cache_invalidate()
    return db_config

@app.get("/api/config", response_model=List[schemas.ConfigOut])
async def read_configs(db: AsyncSession = Depends(get_db)):
    cached = _cache_get("configs")
    if cached is not None:
        return cached
    result = await db.execute(select(models.TimetableConfig))
    data = result.scalars().all()
    _cache_set("configs", data)
    return data

# ═══════════════════════════════════
#  BRANCH  (full CRUD)
# ═══════════════════════════════════
@app.post("/api/branches", response_model=schemas.BranchOut)
async def create_branch(branch: schemas.BranchCreate, db: AsyncSession = Depends(get_db)):
    # Prevent duplicate branch names within the same config
    if branch.config_id is not None:
        existing = await db.execute(
            select(models.Branch).filter(
                models.Branch.config_id == branch.config_id,
                models.Branch.name == branch.name
            )
        )
        if existing.scalars().first():
            raise HTTPException(status_code=400, detail="This entry already exists.")
    db_branch = models.Branch(**branch.model_dump())
    db.add(db_branch)
    await db.commit()
    await db.refresh(db_branch)
    _cache_invalidate()
    return db_branch

@app.get("/api/branches", response_model=List[schemas.BranchOut])
async def read_branches(config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"branches:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Branch)
    if config_id is not None:
        query = query.filter(models.Branch.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

@app.put("/api/branches/{branch_id}", response_model=schemas.BranchOut)
async def update_branch(branch_id: int, data: schemas.BranchUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Branch).filter(models.Branch.id == branch_id))
    branch = result.scalars().first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(branch, k, v)
    await db.commit()
    await db.refresh(branch)
    _cache_invalidate()
    return branch

@app.delete("/api/branches/{branch_id}")
async def delete_branch(branch_id: int, db: AsyncSession = Depends(get_db)):
    # 1. Check if Branch exists
    result = await db.execute(select(models.Branch).filter(models.Branch.id == branch_id))
    branch = result.scalars().first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")

    # 2. Find all Semesters under this branch
    sems_result = await db.execute(select(models.Semester).filter(models.Semester.branch_id == branch_id))
    sems = sems_result.scalars().all()
    sem_ids = [s.id for s in sems]

    # 3. Check for active allocations belonging to these semesters
    if sem_ids:
        alloc_result = await db.execute(select(models.Allocation).filter(models.Allocation.semester_id.in_(sem_ids)))
        if alloc_result.scalars().first():
            raise HTTPException(status_code=400, detail="Warning: Cannot delete branch while related timetable data exists.")

        # 4. Cascade delete manually (SQLite doesn't always handle cascades strictly based on pragmas)
        # Delete Mappings
        await db.execute(sa_delete(models.SemesterFacultyMap).where(models.SemesterFacultyMap.semester_id.in_(sem_ids)))
        await db.execute(sa_delete(models.SemesterRoomMap).where(models.SemesterRoomMap.semester_id.in_(sem_ids)))
        # Delete Subjects
        await db.execute(sa_delete(models.Subject).where(models.Subject.semester_id.in_(sem_ids)))
        # Delete Semesters
        await db.execute(sa_delete(models.Semester).where(models.Semester.branch_id == branch_id))

    # 5. Delete Branch
    await db.delete(branch)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

# ═══════════════════════════════════
#  SEMESTER  (full CRUD)
# ═══════════════════════════════════
@app.post("/api/semesters", response_model=schemas.SemesterOut)
async def create_semester(semester: schemas.SemesterCreate, db: AsyncSession = Depends(get_db)):
    db_semester = models.Semester(**semester.model_dump())
    db.add(db_semester)
    await db.commit()
    await db.refresh(db_semester)
    _cache_invalidate()
    return db_semester

@app.get("/api/semesters", response_model=List[schemas.SemesterOut])
async def read_semesters(config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"semesters:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Semester)
    if config_id is not None:
        query = query.filter(models.Semester.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

@app.put("/api/semesters/{semester_id}", response_model=schemas.SemesterOut)
async def update_semester(semester_id: int, data: schemas.SemesterUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Semester).filter(models.Semester.id == semester_id))
    semester = result.scalars().first()
    if not semester:
        raise HTTPException(status_code=404, detail="Semester not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(semester, k, v)
    await db.commit()
    await db.refresh(semester)
    _cache_invalidate()
    return semester

@app.delete("/api/semesters/{semester_id}")
async def delete_semester(semester_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Semester).filter(models.Semester.id == semester_id))
    semester = result.scalars().first()
    if not semester:
        raise HTTPException(status_code=404, detail="Semester not found")
        
    # Check for active allocations
    alloc_result = await db.execute(select(models.Allocation).filter(models.Allocation.semester_id == semester_id))
    if alloc_result.scalars().first():
         raise HTTPException(status_code=400, detail="Warning: Cannot delete semester while related timetable data exists.")

    # Cascade manual delete
    await db.execute(sa_delete(models.SemesterFacultyMap).where(models.SemesterFacultyMap.semester_id == semester_id))
    await db.execute(sa_delete(models.SemesterRoomMap).where(models.SemesterRoomMap.semester_id == semester_id))
    await db.execute(sa_delete(models.Subject).where(models.Subject.semester_id == semester_id))

    await db.delete(semester)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

# ═══════════════════════════════════
#  SUBJECT  (full CRUD + filter)
# ═══════════════════════════════════
@app.post("/api/subjects", response_model=schemas.SubjectOut)
async def create_subject(subject: schemas.SubjectCreate, db: AsyncSession = Depends(get_db)):
    db_subject = models.Subject(**subject.model_dump())
    db.add(db_subject)
    await db.commit()
    await db.refresh(db_subject)
    _cache_invalidate()
    return db_subject

@app.get("/api/subjects", response_model=List[schemas.SubjectOut])
async def read_subjects(semester_id: Optional[int] = Query(None), config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"subjects:{semester_id}:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Subject)
    if semester_id is not None:
        query = query.filter(models.Subject.semester_id == semester_id)
    if config_id is not None:
        query = query.filter(models.Subject.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

@app.put("/api/subjects/{subject_id}", response_model=schemas.SubjectOut)
async def update_subject(subject_id: int, data: schemas.SubjectUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Subject).filter(models.Subject.id == subject_id))
    subject = result.scalars().first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(subject, k, v)
    await db.commit()
    await db.refresh(subject)
    _cache_invalidate()
    return subject

@app.delete("/api/subjects/{subject_id}")
async def delete_subject(subject_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Subject).filter(models.Subject.id == subject_id))
    subject = result.scalars().first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    await db.delete(subject)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

# ═══════════════════════════════════
#  FACULTY  (full CRUD)
# ═══════════════════════════════════
@app.post("/api/faculties", response_model=schemas.FacultyOut)
async def create_faculty(faculty: schemas.FacultyCreate, db: AsyncSession = Depends(get_db)):
    # Prevent duplicate faculty names within the same config
    if faculty.config_id is not None:
        existing = await db.execute(
            select(models.Faculty).filter(
                models.Faculty.config_id == faculty.config_id,
                models.Faculty.name == faculty.name
            )
        )
        if existing.scalars().first():
            raise HTTPException(status_code=400, detail="This entry already exists.")
    db_faculty = models.Faculty(**faculty.model_dump())
    db.add(db_faculty)
    await db.commit()
    await db.refresh(db_faculty)
    _cache_invalidate()
    return db_faculty

@app.get("/api/faculties", response_model=List[schemas.FacultyOut])
async def read_faculties(config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"faculties:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Faculty)
    if config_id is not None:
        query = query.filter(models.Faculty.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

@app.put("/api/faculties/{faculty_id}", response_model=schemas.FacultyOut)
async def update_faculty(faculty_id: int, data: schemas.FacultyUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Faculty).filter(models.Faculty.id == faculty_id))
    faculty = result.scalars().first()
    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(faculty, k, v)
    await db.commit()
    await db.refresh(faculty)
    _cache_invalidate()
    return faculty

@app.delete("/api/faculties/{faculty_id}")
async def delete_faculty(faculty_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Faculty).filter(models.Faculty.id == faculty_id))
    faculty = result.scalars().first()
    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")
    await db.delete(faculty)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

# ═══════════════════════════════════
#  ROOM  (full CRUD)
# ═══════════════════════════════════
@app.post("/api/rooms", response_model=schemas.RoomOut)
async def create_room(room: schemas.RoomCreate, db: AsyncSession = Depends(get_db)):
    # Prevent duplicate room names within the same config
    if room.config_id is not None:
        existing = await db.execute(
            select(models.Room).filter(
                models.Room.config_id == room.config_id,
                models.Room.name == room.name
            )
        )
        if existing.scalars().first():
            raise HTTPException(status_code=400, detail="This entry already exists.")
    db_room = models.Room(**room.model_dump())
    db.add(db_room)
    await db.commit()
    await db.refresh(db_room)
    _cache_invalidate()
    return db_room

@app.get("/api/rooms", response_model=List[schemas.RoomOut])
async def read_rooms(config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"rooms:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Room)
    if config_id is not None:
        query = query.filter(models.Room.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

@app.put("/api/rooms/{room_id}", response_model=schemas.RoomOut)
async def update_room(room_id: int, data: schemas.RoomUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Room).filter(models.Room.id == room_id))
    room = result.scalars().first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(room, k, v)
    await db.commit()
    await db.refresh(room)
    _cache_invalidate()
    return room

@app.delete("/api/rooms/{room_id}")
async def delete_room(room_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Room).filter(models.Room.id == room_id))
    room = result.scalars().first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    await db.delete(room)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

# ═══════════════════════════════════
#  MAPPINGS  (faculty & room ↔ semester)
# ═══════════════════════════════════
from pydantic import BaseModel

class FacultyMapping(BaseModel):
    semester_id: int
    faculty_id: int

@app.post("/api/mappings/faculty")
async def map_faculty(mapping: FacultyMapping, db: AsyncSession = Depends(get_db)):
    # Prevent duplicate mappings
    existing = await db.execute(
        select(models.SemesterFacultyMap).filter(
            models.SemesterFacultyMap.semester_id == mapping.semester_id,
            models.SemesterFacultyMap.faculty_id == mapping.faculty_id,
        )
    )
    if existing.scalars().first():
        return {"status": "already_mapped"}
    db_map = models.SemesterFacultyMap(semester_id=mapping.semester_id, faculty_id=mapping.faculty_id)
    db.add(db_map)
    await db.commit()
    return {"status": "success"}

@app.get("/api/mappings/faculty/{semester_id}")
async def get_mapped_faculties(semester_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.Faculty).join(models.SemesterFacultyMap).filter(models.SemesterFacultyMap.semester_id == semester_id)
    )
    return result.scalars().all()

@app.delete("/api/mappings/faculty/{semester_id}/{faculty_id}")
async def unmap_faculty(semester_id: int, faculty_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.SemesterFacultyMap).filter(
            models.SemesterFacultyMap.semester_id == semester_id,
            models.SemesterFacultyMap.faculty_id == faculty_id,
        )
    )
    mapping = result.scalars().first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    await db.delete(mapping)
    await db.commit()
    return {"status": "deleted"}

class RoomMapping(BaseModel):
    semester_id: int
    room_id: int

@app.post("/api/mappings/room")
async def map_room(mapping: RoomMapping, db: AsyncSession = Depends(get_db)):
    db_map = models.SemesterRoomMap(semester_id=mapping.semester_id, room_id=mapping.room_id)
    db.add(db_map)
    await db.commit()
    return {"status": "success"}

@app.get("/api/mappings/room/{semester_id}")
async def get_mapped_rooms(semester_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.Room).join(models.SemesterRoomMap).filter(models.SemesterRoomMap.semester_id == semester_id)
    )
    return result.scalars().all()

# ═══════════════════════════════════
#  ALLOCATIONS & COLLISION LOGIC
# ═══════════════════════════════════
def add_minutes(t: time, mins: int) -> time:
    dt = datetime.combine(date.today(), t) + timedelta(minutes=mins)
    return dt.time()

def check_overlap(start1: time, end1: time, start2: time, end2: time) -> bool:
    return max(start1, start2) < min(end1, end2)

@app.post("/api/allocations", response_model=schemas.AllocationOut)
async def create_allocation(allocation: schemas.AllocationCreate, db: AsyncSession = Depends(get_db)):
    new_start = allocation.start_time
    new_end = add_minutes(new_start, allocation.duration_minutes)

    result = await db.execute(select(models.Allocation).filter(
        models.Allocation.day_of_week == allocation.day_of_week,
        models.Allocation.config_id == allocation.config_id
    ))
    existing_allocations = result.scalars().all()

    for ext in existing_allocations:
        ext_start = ext.start_time
        ext_end = add_minutes(ext_start, ext.duration_minutes)

        if check_overlap(new_start, new_end, ext_start, ext_end):
            if ext.faculty_id == allocation.faculty_id:
                # Check if this faculty is marked to ignore collisions
                fac_result = await db.execute(select(models.Faculty).filter(models.Faculty.id == allocation.faculty_id))
                faculty = fac_result.scalars().first()
                if not faculty or not faculty.ignore_collision:
                    raise HTTPException(status_code=400, detail="This faculty is already assigned in this time slot.")

    db_allocation = models.Allocation(**allocation.model_dump())
    db.add(db_allocation)
    await db.commit()
    await db.refresh(db_allocation)
    return db_allocation

@app.put("/api/allocations/{allocation_id}", response_model=schemas.AllocationOut)
async def update_allocation(allocation_id: int, data: schemas.AllocationUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Allocation).filter(models.Allocation.id == allocation_id))
    db_allocation = result.scalars().first()
    if not db_allocation:
        raise HTTPException(status_code=404, detail="Allocation not found")
        
    update_data = data.model_dump(exclude_unset=True)
    if not update_data:
        return db_allocation

    # Overlap validation logic if changing subject, faculty, room or duration
    # We validate using the proposed state of this allocation
    new_day = db_allocation.day_of_week
    new_start = db_allocation.start_time
    # Use existing or proposed details
    prop_duration = update_data.get('duration_minutes', db_allocation.duration_minutes)
    prop_faculty = update_data.get('faculty_id', db_allocation.faculty_id)
    prop_sem = db_allocation.semester_id # Fixed in this context
    prop_batches = update_data.get('batches', db_allocation.batches)
    
    new_end = add_minutes(new_start, prop_duration)

    ext_result = await db.execute(select(models.Allocation).filter(
        models.Allocation.day_of_week == new_day,
        models.Allocation.config_id == db_allocation.config_id,
        models.Allocation.id != allocation_id
    ))
    existing_allocations = ext_result.scalars().all()

    for ext in existing_allocations:
        ext_start = ext.start_time
        ext_end = add_minutes(ext_start, ext.duration_minutes)

        if check_overlap(new_start, new_end, ext_start, ext_end):
            if ext.faculty_id == prop_faculty:
                # Check if this faculty is marked to ignore collisions
                fac_result = await db.execute(select(models.Faculty).filter(models.Faculty.id == prop_faculty))
                faculty = fac_result.scalars().first()
                if not faculty or not faculty.ignore_collision:
                    raise HTTPException(status_code=400, detail="This faculty is already assigned in this time slot.")

    for k, v in update_data.items():
        setattr(db_allocation, k, v)
        
    await db.commit()
    await db.refresh(db_allocation)
    _cache_invalidate()
    return db_allocation

@app.delete("/api/allocations/{allocation_id}")
async def delete_allocation(allocation_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Allocation).filter(models.Allocation.id == allocation_id))
    allocation = result.scalars().first()
    if not allocation:
        raise HTTPException(status_code=404, detail="Allocation not found")
    await db.delete(allocation)
    await db.commit()
    _cache_invalidate()
    return {"status": "deleted"}

@app.get("/api/allocations", response_model=List[schemas.AllocationOut])
async def read_allocations(config_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    cache_key = f"allocations:{config_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    query = select(models.Allocation)
    if config_id is not None:
        query = query.filter(models.Allocation.config_id == config_id)
    result = await db.execute(query)
    data = result.scalars().all()
    _cache_set(cache_key, data)
    return data

# ═══════════════════════════════════
#  EXPORT
# ═══════════════════════════════════
@app.get("/api/export")
async def get_export_data(db: AsyncSession = Depends(get_db)):
    allocations = (await db.execute(select(models.Allocation))).scalars().all()
    return allocations

# ─── openpyxl-based GTU-format Excel export ───
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

DAYS_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

def _time_to_minutes(t) -> int:
    """Convert a time object or HH:MM:SS string to minutes since midnight."""
    if isinstance(t, str):
        parts = t.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    return t.hour * 60 + t.minute

def _mins_to_hhmm(mins: int) -> str:
    return f"{mins // 60:02d}:{mins % 60:02d}"

def _generate_timeslots(config) -> list:
    """Generate timeslots with breaks, matching the frontend logic."""
    slots = []
    breaks_raw = config.breaks or []
    start_mins = _time_to_minutes(config.start_time)
    end_mins = _time_to_minutes(config.end_time)
    slot_dur = config.slot_duration_minutes
    cursor = start_mins

    # Parse break start times
    break_map = {}
    for b in breaks_raw:
        bstart = b.get('start_time', '') if isinstance(b, dict) else ''
        bdur = b.get('duration_minutes', 0) if isinstance(b, dict) else 0
        if bstart:
            break_map[_time_to_minutes(bstart)] = bdur

    while cursor < end_mins:
        if cursor in break_map:
            bdur = break_map[cursor]
            slots.append({'type': 'break', 'start_mins': cursor, 'end_mins': cursor + bdur, 'display': 'RECESS'})
            cursor += bdur
        else:
            slot_end = cursor + slot_dur
            slots.append({
                'type': 'slot',
                'start_mins': cursor,
                'end_mins': slot_end,
                'display': f"{_mins_to_hhmm(cursor)}-{_mins_to_hhmm(slot_end)}"
            })
            cursor = slot_end
    return slots

def _build_sheet(ws, title: str, allocs, timeslots, all_subjects, all_faculties, all_rooms, slot_duration: int):
    """Build a single GTU-format sheet."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    total_cols = 2 + len(DAYS_ORDER)  # Lec No + Time + 6 days
    last_col_letter = get_column_letter(total_cols)

    # Row 1: University
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'Gujarat Technological University'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center

    # Row 2: School
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'SCHOOL OF ENGINEERING AND TECHNOLOGY'
    ws['A2'].font = Font(bold=True, size=13)
    ws['A2'].alignment = center

    # Row 3: Title
    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'] = f'Class Timetable: {title}'
    ws['A3'].font = title_font
    ws['A3'].alignment = center

    # Row 4: Coordinator
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'] = 'Class Coordinator: __________'
    ws['A4'].font = Font(size=11)
    ws['A4'].alignment = left_align

    # Row 5: Header
    header_row = 5
    headers = ['Lec No', 'Time'] + DAYS_ORDER
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Build lookup: (day, start_mins) -> list of allocations
    alloc_map: dict = {}
    for a in allocs:
        a_start = _time_to_minutes(a.start_time)
        key = (a.day_of_week, a_start)
        alloc_map.setdefault(key, []).append(a)

    # Subject/faculty/room name lookups
    sub_map = {s.id: s.name for s in all_subjects}
    fac_map = {f.id: f.name for f in all_faculties}
    rm_map = {r.id: r.name for r in all_rooms}

    # Track which rows to merge for multi-slot labs (per day column)
    # We'll first write all data, then merge afterwards
    data_start_row = header_row + 1
    current_row = data_start_row
    lec_no = 0

    # Track merge ranges for labs: list of (start_row, end_row, col)
    merge_ranges = []

    # Map from (slot_index) to row number for merge tracking
    slot_row_map = {}

    for si, slot in enumerate(timeslots):
        row = current_row
        slot_row_map[si] = row

        if slot['type'] == 'break':
            # Break row
            ws.cell(row=row, column=1, value='Break').font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).border = border

            ws.cell(row=row, column=2, value=slot['display']).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).font = Font(bold=True, size=11)

            # Merge day columns with RECESS
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_cols)
            ws.cell(row=row, column=3, value='RECESS')
            ws.cell(row=row, column=3).alignment = center
            ws.cell(row=row, column=3).font = Font(bold=True, size=12, color='FF0000')
            ws.cell(row=row, column=3).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=3).border = border
        else:
            lec_no += 1
            ws.cell(row=row, column=1, value=lec_no).alignment = center
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)

            ws.cell(row=row, column=2, value=slot['display']).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).font = Font(size=10)

            for di, day in enumerate(DAYS_ORDER):
                col = 3 + di
                cell_allocs = alloc_map.get((day, slot['start_mins']), [])

                if cell_allocs:
                    lines = []
                    for a in cell_allocs:
                        sub_name = sub_map.get(a.subject_id, '?')
                        fac_name = fac_map.get(a.faculty_id, '?')
                        rm_name = rm_map.get(a.room_id, '?')
                        batches = a.batches or []

                        if batches:
                            for batch in batches:
                                lines.append(f"{batch}\n{sub_name}({fac_name})({rm_name})")
                        else:
                            lines.append(f"{sub_name}({fac_name})({rm_name})")

                        # Check if this is a multi-slot lab
                        if a.duration_minutes > slot_duration:
                            spans = a.duration_minutes // slot_duration
                            end_slot_idx = si + spans - 1
                            if end_slot_idx < len(timeslots):
                                end_row = slot_row_map.get(si, row) + spans - 1
                                merge_ranges.append((row, end_row, col))

                    cell_value = '\n'.join(lines)
                    cell = ws.cell(row=row, column=col, value=cell_value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                    cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=row, column=col, value='')
                    cell.border = border

        current_row += 1

    # Apply lab merges
    for start_r, end_r, col in merge_ranges:
        if end_r > start_r and end_r < current_row:
            try:
                ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
            except Exception:
                pass  # Skip if already merged or overlapping

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # Set row heights
    for r in range(data_start_row, current_row):
        ws.row_dimensions[r].height = 45


@app.get("/api/export_excel")
async def export_excel(
    config_id: int = Query(...),
    mode: str = Query('all'),   # master, selected, all
    value: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    # Fetch config
    config_result = await db.execute(select(models.TimetableConfig).filter(models.TimetableConfig.id == config_id))
    config = config_result.scalars().first()
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")

    # Fetch all reference data for this config
    all_branches = (await db.execute(select(models.Branch).filter(models.Branch.config_id == config_id))).scalars().all()
    all_semesters = (await db.execute(select(models.Semester).filter(models.Semester.config_id == config_id))).scalars().all()
    all_subjects = (await db.execute(select(models.Subject).filter(models.Subject.config_id == config_id))).scalars().all()
    all_faculties = (await db.execute(select(models.Faculty).filter(models.Faculty.config_id == config_id))).scalars().all()
    all_rooms = (await db.execute(select(models.Room).filter(models.Room.config_id == config_id))).scalars().all()
    all_allocations = (await db.execute(select(models.Allocation).filter(models.Allocation.config_id == config_id))).scalars().all()

    timeslots = _generate_timeslots(config)
    branch_map = {b.id: b.name for b in all_branches}
    sem_map = {s.id: s for s in all_semesters}

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    filename = f"{config.name or 'Timetable'}.xlsx"

    if mode == 'master':
        ws = wb.create_sheet(title="Master Timetable"[:31])
        # For master timetable, we want ONLY master allocs, or all the slots? 
        # The prompt says: Allocation.query.filter_by(is_master=True)
        # But wait! We do not have an is_master flag right now on the actual code, 
        # config isolation handles "Master" vs "Departmental" already for the user if they've designed it that way. 
        # We'll just pass all_allocations directly as we did before.
        _build_sheet(ws, "Master Timetable", all_allocations, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)
        filename = "Master_Timetable.xlsx"

    elif mode == 'selected' and value:
        try:
            item_type, item_id_str = value.split(':')
            item_id = int(item_id_str)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid value format")

        if item_type == 'semester':
            sem = sem_map.get(item_id)
            if sem:
                branch_name = branch_map.get(sem.branch_id, '')
                sheet_title = f"{branch_name} {sem.name}".strip()
                ws = wb.create_sheet(title=sheet_title[:31])
                allocs = [a for a in all_allocations if a.semester_id == item_id]
                _build_sheet(ws, sheet_title, allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)
                filename = f"{branch_name}_{sem.name}_Timetable.xlsx".replace(' ', '_')

        elif item_type == 'faculty':
            fac = next((f for f in all_faculties if f.id == item_id), None)
            if fac:
                ws = wb.create_sheet(title=f"Faculty_{fac.name}"[:31])
                allocs = [a for a in all_allocations if a.faculty_id == item_id]
                _build_sheet(ws, f"Faculty: {fac.name}", allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)
                filename = f"Faculty_{fac.name}_Timetable.xlsx".replace(' ', '_')

        elif item_type == 'room':
            rm = next((r for r in all_rooms if r.id == item_id), None)
            if rm:
                ws = wb.create_sheet(title=f"Room_{rm.name}"[:31])
                allocs = [a for a in all_allocations if a.room_id == item_id]
                _build_sheet(ws, f"Room: {rm.name}", allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)
                filename = f"Room_{rm.name}_Timetable.xlsx".replace(' ', '_')

    elif mode == 'all':
        # Master sheet first
        ws_master = wb.create_sheet(title="Master Timetable"[:31])
        _build_sheet(ws_master, "Master Timetable", all_allocations, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)

        # Branch sheets
        for sem in all_semesters:
            branch_name = branch_map.get(sem.branch_id, '')
            sheet_title = f"{branch_name} {sem.name}".strip()[:31]
            ws = wb.create_sheet(title=sheet_title)
            allocs = [a for a in all_allocations if a.semester_id == sem.id]
            _build_sheet(ws, f"{branch_name} ({sem.name})", allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)

        # Faculty sheets
        for fac in all_faculties:
            fac_allocs = [a for a in all_allocations if a.faculty_id == fac.id]
            if fac_allocs:
                ws = wb.create_sheet(title=f"Fac_{fac.name}"[:31])
                _build_sheet(ws, f"Faculty: {fac.name}", fac_allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)

        # Room sheets
        for rm in all_rooms:
            rm_allocs = [a for a in all_allocations if a.room_id == rm.id]
            if rm_allocs:
                ws = wb.create_sheet(title=f"Room_{rm.name}"[:31])
                _build_sheet(ws, f"Room: {rm.name}", rm_allocs, timeslots, all_subjects, all_faculties, all_rooms, config.slot_duration_minutes)

        filename = "All_Timetables.xlsx"

    # Fallback if no sheets were created
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Empty")
        ws['A1'] = 'No data to export'

    # Write to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
